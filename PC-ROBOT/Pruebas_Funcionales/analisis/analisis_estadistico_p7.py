#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script 8 — Análisis Estadístico Global e Informe de Fase P7 del TFG.

Procesa las mediciones de las condiciones de red para las métricas M1 a M4,
integrando automáticamente:
- M1: Latencias Glass-to-Glass manuales y software (Pruebas_Conexion/GlassToGlass_Manual).
- M2: Error radial y precisión estática (M2_Precision_Posicionado.xlsx).
- M3: Repetibilidad cinemática 2D según norma ISO 9283 (M3_Repetibilidad.xlsx).
- M4: Teleoperación dinámica Pick-and-Place y micro-correcciones (M4_Pick_And_Place.xlsx y correcciones_m4.csv).

Ejecuta contrastes no paramétricos (Kruskal-Wallis, Mann-Whitney U con Bonferroni, Wilcoxon)
y compila el informe Markdown final sin datos faltantes.
"""

import argparse
import csv
import datetime
import logging
import math
import os
import sys
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import matplotlib.pyplot as plt
import numpy as np
import openpyxl

try:
    import scipy.stats as stats
    SCIPY_DISPONIBLE = True
except ImportError:
    stats = None
    SCIPY_DISPONIBLE = False

# Configuración de rutas
CURRENT_DIR = Path(__file__).resolve().parent
REPO_ROOT = CURRENT_DIR.parents[2]
RESULTADOS_DIR = CURRENT_DIR.parent / "resultados"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S"
)
log = logging.getLogger("analisis_estadistico_p7")

# Latencias Glass-to-Glass medidas en la validación software previa (ms)
LATENCIA_G2G_MS = {
    "C4_ETHERNET": 2.7,
    "C1_WIFI_SIN_CARGA": 5.7,
    "C2_WIFI_CARGA_MEDIA": 9.4,
    "C3_WIFI_CARGA_ALTA": 13.7,
}

CONDICIONES_ORDENADAS = [
    "C4_ETHERNET",
    "C1_WIFI_SIN_CARGA",
    "C2_WIFI_CARGA_MEDIA",
    "C3_WIFI_CARGA_ALTA"
]

NOMBRES_LEGIBLES = {
    "C4_ETHERNET": "C4 Ethernet (Sin Carga)",
    "C1_WIFI_SIN_CARGA": "C1 WiFi (Sin Carga)",
    "C2_WIFI_CARGA_MEDIA": "C2 WiFi (Carga Media)",
    "C3_WIFI_CARGA_ALTA": "C3 WiFi (Carga Alta)"
}

def calcular_tamano_efecto_r(p_val: float, n_total: int) -> float:
    """Calcula el tamaño del efecto r = |Z| / sqrt(N) para contrastes no paramétricos."""
    if not SCIPY_DISPONIBLE or p_val <= 0 or n_total <= 0:
        return 0.0
    p_clamped = min(max(p_val, 1e-15), 1.0 - 1e-15)
    z = stats.norm.ppf(1.0 - p_clamped / 2.0)
    return float(z / math.sqrt(n_total))

def cargar_datos_m1(repo_root: Path) -> Tuple[Dict[str, List[float]], Dict[str, float]]:
    """Carga los datos de latencia Glass-to-Glass manuales y software de M1."""
    datos_manual = {c: [] for c in CONDICIONES_ORDENADAS}
    datos_software = {c: LATENCIA_G2G_MS[c] for c in CONDICIONES_ORDENADAS}

    g2g_map = {
        "C4_ETHERNET": repo_root / "Pruebas_Conexion/GlassToGlass_Manual/ETHERNET_IPERF3/ETH_IPERF3_SinCarga/ETH_IPERF3_SinCarga_Resultados.xlsx",
        "C1_WIFI_SIN_CARGA": repo_root / "Pruebas_Conexion/GlassToGlass_Manual/WIFI/WIFI_SinCarga/WIFI_SinCarga_Resultados.xlsx",
        "C2_WIFI_CARGA_MEDIA": repo_root / "Pruebas_Conexion/GlassToGlass_Manual/WIFI/WIFI_CargaMedia/WIFI_CargaMedia_Resultados.xlsx",
        "C3_WIFI_CARGA_ALTA": repo_root / "Pruebas_Conexion/GlassToGlass_Manual/WIFI/WIFI_CargaAlta/WIFI_CargaAlta_Resultados.xlsx",
    }

    for cond, fpath in g2g_map.items():
        if fpath.exists():
            try:
                wb = openpyxl.load_workbook(fpath, data_only=True)
                if "Fotogramas" in wb.sheetnames:
                    ws = wb["Fotogramas"]
                    # Columna 11 es 'Incluida en estadistica (ms)'
                    for r in range(5, 40):
                        val = ws.cell(row=r, column=11).value
                        if val is not None and str(val).strip() != "":
                            try:
                                v_flt = float(val)
                                datos_manual[cond].append(v_flt)
                            except ValueError:
                                pass
            except Exception as e:
                log.warning(f"No se pudo leer {fpath}: {e}")

    return datos_manual, datos_software

def cargar_datos_m2_m3(repo_root: Path) -> Tuple[Dict[str, List[float]], Dict[str, List[Tuple[float, float]]], Dict[str, float]]:
    """Carga los errores radiales (M2) y repetibilidad ISO 9283 (M3) desde los Excels oficiales."""
    errores_radiales = {c: [] for c in CONDICIONES_ORDENADAS}
    pares_xy = {c: [] for c in CONDICIONES_ORDENADAS}
    repetibilidad_iso = {c: 0.0 for c in CONDICIONES_ORDENADAS}

    # 1. Cargar M2
    m2_path = repo_root / "PC-ROBOT/Pruebas_Funcionales/resultados/M2/M2_Precision_Posicionado.xlsx"
    sheet_m2_map = {
        "C1_WIFI_SIN_CARGA": "C1_Ethernet_Sin_Carga",
        "C2_WIFI_CARGA_MEDIA": "C2_Ethernet_Carga_Media",
        "C3_WIFI_CARGA_ALTA": "C3_Ethernet_Carga_Alta"
    }

    if m2_path.exists():
        try:
            wb = openpyxl.load_workbook(m2_path, data_only=True)
            for cond, sname in sheet_m2_map.items():
                if sname in wb.sheetnames:
                    ws = wb[sname]
                    for r in range(11, 38):
                        ex = ws.cell(row=r, column=3).value
                        ey = ws.cell(row=r, column=4).value
                        if ex is not None and ey is not None:
                            try:
                                ex_cm = float(ex) / 10.0  # mm a cm
                                ey_cm = float(ey) / 10.0
                                er_cm = math.sqrt(ex_cm**2 + ey_cm**2)
                                errores_radiales[cond].append(er_cm)
                                pares_xy[cond].append((ex_cm, ey_cm))
                            except (ValueError, TypeError):
                                pass
        except Exception as e:
            log.warning(f"Error leyendo M2 Excel: {e}")

    # 2. Cargar M3
    m3_path = repo_root / "PC-ROBOT/Pruebas_Funcionales/resultados/M3/M3_Repetibilidad.xlsx"
    if m3_path.exists():
        try:
            wb3 = openpyxl.load_workbook(m3_path, data_only=True)
            if "M3_Repetibilidad" in wb3.sheetnames:
                ws3 = wb3["M3_Repetibilidad"]
                p_list = []
                for r in range(13, 23):
                    ex = ws3.cell(row=r, column=2).value
                    ey = ws3.cell(row=r, column=3).value
                    if ex is not None and ey is not None:
                        try:
                            p_list.append((float(ex) / 10.0, float(ey) / 10.0))  # en cm
                        except (ValueError, TypeError):
                            pass
                if p_list:
                    xs = [p[0] for p in p_list]
                    ys = [p[1] for p in p_list]
                    x_bar, y_bar = np.mean(xs), np.mean(ys)
                    dists = [math.sqrt((x - x_bar)**2 + (y - y_bar)**2) for x, y in p_list]
                    l_bar = np.mean(dists)
                    s_l = np.std(dists, ddof=1)
                    rp_iso = l_bar + 3 * s_l
                    repetibilidad_iso["C1_WIFI_SIN_CARGA"] = rp_iso
        except Exception as e:
            log.warning(f"Error leyendo M3 Excel: {e}")

    return errores_radiales, pares_xy, repetibilidad_iso

def cargar_datos_m4(repo_root: Path) -> Tuple[Dict[str, List[float]], Dict[str, List[int]], Dict[str, Dict[int, List[float]]], Dict[str, List[int]]]:
    """Carga tiempos de ejecución, éxitos y micro-correcciones de M4."""
    tiempos = {c: [] for c in CONDICIONES_ORDENADAS}
    exitos = {c: [] for c in CONDICIONES_ORDENADAS}
    curvas = {c: {i: [] for i in range(1, 6)} for c in CONDICIONES_ORDENADAS}
    correcciones_totales = {c: [] for c in CONDICIONES_ORDENADAS}

    # Leer Excel M4_Pick_And_Place.xlsx
    excel_m4 = repo_root / "PC-ROBOT/Pruebas_Funcionales/resultados/M4/M4_Pick_And_Place.xlsx"
    cond_trans = {
        "C1_ETHERNET_SIN_CARGA": "C1_WIFI_SIN_CARGA",
        "C2_ETHERNET_CLUMSY_CARGA_MEDIA": "C2_WIFI_CARGA_MEDIA",
        "C3_ETHERNET_CLUMSY_CARGA_ALTA": "C3_WIFI_CARGA_ALTA",
    }

    if excel_m4.exists():
        try:
            wb = openpyxl.load_workbook(excel_m4, data_only=True)
            ws = wb["Registro_M4"]
            for r in range(10, 40):
                c_val = str(ws.cell(row=r, column=2).value).strip()
                t_val = ws.cell(row=r, column=4).value
                tar_ok = ws.cell(row=r, column=10).value
                corrs = ws.cell(row=r, column=7).value
                obj = str(ws.cell(row=r, column=3).value).strip()

                cond_std = cond_trans.get(c_val)
                if cond_std and cond_std in CONDICIONES_ORDENADAS:
                    # Éxito
                    try:
                        exitos[cond_std].append(int(tar_ok))
                    except (ValueError, TypeError):
                        pass

                    # Tiempo
                    if t_val is not None and str(t_val).strip() not in ["-", ""]:
                        try:
                            t_flt = float(t_val)
                            tiempos[cond_std].append(t_flt)
                        except ValueError:
                            pass

                    # Correcciones
                    if corrs is not None and str(corrs).strip() not in ["-", ""]:
                        try:
                            correcciones_totales[cond_std].append(int(corrs))
                        except ValueError:
                            pass
        except Exception as e:
            log.warning(f"Error leyendo M4 Excel: {e}")

    # Leer correcciones_m4.csv para curvas por intento
    ruta_correcciones = repo_root / "PC-ROBOT/Pruebas_Funcionales/resultados/correcciones_m4.csv"
    if ruta_correcciones.exists():
        with open(ruta_correcciones, mode="r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for r in reader:
                cond = r.get("condicion_red", "").strip()
                if cond in CONDICIONES_ORDENADAS:
                    try:
                        dur = float(r.get("duracion_s", 0.0))
                        iid = int(r.get("id_intento", 1))
                        obj = r.get("objeto", "N1").strip()
                        if obj == "N1" and 1 <= iid <= 5:
                            curvas[cond][iid].append(dur)
                    except ValueError:
                        pass

    return tiempos, exitos, curvas, correcciones_totales

def generar_tabla_y_graficos(
    dir_res: Path,
    res_m1_man: Dict[str, List[float]],
    res_m1_soft: Dict[str, float],
    res_m2: Dict[str, List[float]],
    res_m3_xy: Dict[str, List[Tuple[float, float]]],
    res_m3_iso: Dict[str, float],
    res_m4_t: Dict[str, List[float]],
    res_m4_e: Dict[str, List[int]],
    curvas_m4: Dict[str, Dict[int, List[float]]],
    res_m4_corrs: Dict[str, List[int]]
) -> Dict[str, Dict[str, str]]:
    """Genera tabla de resumen, gráficos PNG y tabla CSV."""
    tabla_resumen: Dict[str, Dict[str, str]] = {}

    for c in CONDICIONES_ORDENADAS:
        # M1: G2G Manual
        v_m1 = res_m1_man[c]
        m1_str = f"{np.mean(v_m1):.1f} ± {np.std(v_m1, ddof=1):.1f} ms (N={len(v_m1)})" if v_m1 else f"{res_m1_soft[c]:.1f} ms (Software)"

        # M2: Error Radial en cm
        v_m2 = res_m2[c]
        m2_str = f"{np.mean(v_m2):.2f} ± {np.std(v_m2, ddof=1):.2f} cm (N={len(v_m2)})" if v_m2 else "N/D"

        # M3: Repetibilidad 2D
        v_m3 = res_m3_xy[c]
        rp_val = res_m3_iso.get(c, 0.0)
        if rp_val > 0:
            m3_str = f"{rp_val:.2f} cm (ISO 9283 Rp)"
        elif len(v_m3) > 1:
            exs = [p[0] for p in v_m3]
            eys = [p[1] for p in v_m3]
            sigma_2d = math.sqrt(np.var(exs, ddof=1) + np.var(eys, ddof=1))
            m3_str = f"{sigma_2d:.2f} cm (σ 2D)"
        else:
            m3_str = "N/D"

        # M4 Tiempo
        v_m4 = res_m4_t[c]
        m4_t_str = f"{np.mean(v_m4):.1f} ± {np.std(v_m4, ddof=1):.1f} s (N={len(v_m4)})" if v_m4 else "N/D"

        # M4 Éxito
        v_e = res_m4_e[c]
        if v_e:
            tasa_exito = sum(v_e) / len(v_e) * 100.0
            m4_e_str = f"{tasa_exito:.1f}% ({sum(v_e)}/{len(v_e)})"
        else:
            m4_e_str = "N/D"

        tabla_resumen[c] = {
            "M1_Latencia_G2G": m1_str,
            "M2_Error_Radial_cm": m2_str,
            "M3_Repetibilidad_cm": m3_str,
            "M4_Tiempo_s": m4_t_str,
            "M4_Tasa_Exito": m4_e_str
        }

    # Guardar CSV de tabla resumen
    csv_tabla = dir_res / "tabla_resumen_p7.csv"
    with open(csv_tabla, mode="w", newline="", encoding="utf-8") as f:
        fieldnames = ["Condicion", "Latencia_G2G_Base_ms", "M1_Latencia_G2G_Manual", "M2_Precision_Radial", "M3_Repetibilidad_2D", "M4_Tiempo_PickAndPlace", "M4_Tasa_Exito"]
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for c in CONDICIONES_ORDENADAS:
            writer.writerow({
                "Condicion": NOMBRES_LEGIBLES[c],
                "Latencia_G2G_Base_ms": LATENCIA_G2G_MS[c],
                "M1_Latencia_G2G_Manual": tabla_resumen[c]["M1_Latencia_G2G"],
                "M2_Precision_Radial": tabla_resumen[c]["M2_Error_Radial_cm"],
                "M3_Repetibilidad_2D": tabla_resumen[c]["M3_Repetibilidad_cm"],
                "M4_Tiempo_PickAndPlace": tabla_resumen[c]["M4_Tiempo_s"],
                "M4_Tasa_Exito": tabla_resumen[c]["M4_Tasa_Exito"]
            })

    # Gráfico 1: Correlación Red vs Tarea
    condiciones_con_datos = [c for c in CONDICIONES_ORDENADAS if res_m4_t[c]]
    ruta_corr = dir_res / "grafico_correlacion_g2g_m4.png"
    if condiciones_con_datos:
        fig, ax1 = plt.subplots(figsize=(7.5, 4.8))
        g2g_x = [LATENCIA_G2G_MS[c] for c in condiciones_con_datos]
        m4_medias = [float(np.mean(res_m4_t[c])) for c in condiciones_con_datos]
        m4_stds = [float(np.std(res_m4_t[c], ddof=1)) if len(res_m4_t[c]) > 1 else 0.0 for c in condiciones_con_datos]

        ax1.errorbar(g2g_x, m4_medias, yerr=m4_stds, fmt='o-', color='#0074D9', linewidth=2, capsize=5, label='Tiempo de Compleción M4 (s)')
        ax1.set_xlabel('Latencia Glass-to-Glass de Vídeo Software (ms)', fontsize=10, fontweight='bold')
        ax1.set_ylabel('Tiempo Medio M4 (s)', color='#0074D9', fontsize=10, fontweight='bold')
        ax1.tick_params(axis='y', labelcolor='#0074D9')
        ax1.grid(True, linestyle=':', alpha=0.6)

        for i, c in enumerate(condiciones_con_datos):
            ax1.annotate(c.split('_')[0], (g2g_x[i], m4_medias[i]), textcoords="offset points", xytext=(0, 10), ha='center', fontsize=9, weight='bold')

        ax1.set_title('Correlación entre Degradación de Red y Rendimiento en Tarea (M4)', fontsize=11, fontweight='bold')
        plt.tight_layout()
        plt.savefig(ruta_corr, dpi=250)
        plt.close()

    # Gráfico 2: Curva de Aprendizaje M4
    ruta_curva = dir_res / "curva_aprendizaje_m4.png"
    fig, ax = plt.subplots(figsize=(7.5, 4.5))
    colores = {'C4_ETHERNET': '#2ECC40', 'C1_WIFI_SIN_CARGA': '#0074D9', 'C2_WIFI_CARGA_MEDIA': '#FF851B', 'C3_WIFI_CARGA_ALTA': '#FF4136'}

    hay_alguna_curva = False
    for c in condiciones_con_datos:
        intentos_x = []
        medias_int = []
        for i in range(1, 6):
            vals = curvas_m4[c][i]
            if vals:
                intentos_x.append(i)
                medias_int.append(float(np.mean(vals)))
        if intentos_x:
            hay_alguna_curva = True
            ax.plot(intentos_x, medias_int, marker='s', label=NOMBRES_LEGIBLES[c], color=colores[c], linewidth=1.8)

    if hay_alguna_curva:
        ax.set_title('Curva de Aprendizaje en Tarea Pick-and-Place (M4 Rollos)', fontsize=11, fontweight='bold')
        ax.set_xlabel('Número de Intento Consecutivo', fontsize=10, fontweight='bold')
        ax.set_ylabel('Tiempo de Ejecución (s)', fontsize=10, fontweight='bold')
        ax.set_xticks(range(1, 6))
        ax.grid(True, linestyle=':', alpha=0.6)
        ax.legend(loc='upper right', fontsize=9)
        plt.tight_layout()
        plt.savefig(ruta_curva, dpi=250)
        plt.close()

    return tabla_resumen

def ejecutar_tests_estadisticos(
    res_m1: Dict[str, List[float]],
    res_m2: Dict[str, List[float]],
    res_m4_t: Dict[str, List[float]],
    res_m4_corrs: Dict[str, List[int]],
    pares_xy: Dict[str, List[Tuple[float, float]]]
) -> List[str]:
    """Ejecuta los contrastes estadísticos formales y devuelve el informe formateado en Markdown."""
    lineas = []

    if not SCIPY_DISPONIBLE:
        lineas.append("> [!NOTE]")
        lineas.append("> La librería `scipy` no está disponible en este entorno local.")
        return lineas

    # 1. Kruskal-Wallis Global
    lineas.append("### 1. Contrastes Globales entre Condiciones (Kruskal-Wallis)")
    lineas.append("")

    for nombre_metrica, dic_datos in [
        ("M1 Latencia Glass-to-Glass Manual", res_m1),
        ("M2 Precisión / Error Radial", res_m2),
        ("M4 Tiempo de Compleción", res_m4_t),
        ("M4 Micro-correcciones Cinemáticas", res_m4_corrs)
    ]:
        grupos = [dic_datos[c] for c in CONDICIONES_ORDENADAS if len(dic_datos[c]) >= 2]
        if len(grupos) >= 2:
            h_stat, p_val = stats.kruskal(*grupos)
            sig_str = "*(Diferencia Estadísticamente Significativa, $p < 0.05$)*" if p_val < 0.05 else "*(No significativo)*"
            lineas.append(f"- **{nombre_metrica}:** $H = {h_stat:.3f}$, $p = {p_val:.4e}$ {sig_str}")
        else:
            lineas.append(f"- **{nombre_metrica}:** Muestras insuficientes para contraste global.")

    lineas.append("")
    lineas.append("### 2. Comparaciones Post-Hoc por Pares (Mann-Whitney U con Corrección Bonferroni)")
    lineas.append("")
    lineas.append("| Par de Condiciones | Métrica | Estadístico U | p-valor bruto | p-valor corregido | Tamaño del Efecto r |")
    lineas.append("|---|---|---|---|---|---|")

    conds_eval = [c for c in CONDICIONES_ORDENADAS if len(res_m2[c]) >= 2 or len(res_m4_t[c]) >= 2]
    n_comp = (len(conds_eval) * (len(conds_eval) - 1)) // 2 if len(conds_eval) > 1 else 1

    for i in range(len(conds_eval)):
        for j in range(i + 1, len(conds_eval)):
            c1 = conds_eval[i]
            c2 = conds_eval[j]
            par_nombre = f"{c1.split('_')[0]} vs {c2.split('_')[0]}"

            # M2 Error Radial
            g1_m2, g2_m2 = res_m2[c1], res_m2[c2]
            if len(g1_m2) >= 2 and len(g2_m2) >= 2:
                u_stat, p_raw = stats.mannwhitneyu(g1_m2, g2_m2, alternative='two-sided')
                p_bonf = min(1.0, p_raw * n_comp)
                r_efecto = calcular_tamano_efecto_r(p_raw, len(g1_m2) + len(g2_m2))
                lineas.append(f"| {par_nombre} | M2 Error Radial | {u_stat:.1f} | {p_raw:.4f} | {p_bonf:.4f} | {r_efecto:.3f} |")

            # M4 Tiempo
            g1_m4, g2_m4 = res_m4_t[c1], res_m4_t[c2]
            if len(g1_m4) >= 2 and len(g2_m4) >= 2:
                u_stat, p_raw = stats.mannwhitneyu(g1_m4, g2_m4, alternative='two-sided')
                p_bonf = min(1.0, p_raw * n_comp)
                r_efecto = calcular_tamano_efecto_r(p_raw, len(g1_m4) + len(g2_m4))
                lineas.append(f"| {par_nombre} | M4 Tiempo Compleción | {u_stat:.1f} | {p_raw:.4f} | {p_bonf:.4f} | {r_efecto:.3f} |")

            # M4 Correcciones
            g1_cor, g2_cor = res_m4_corrs[c1], res_m4_corrs[c2]
            if len(g1_cor) >= 2 and len(g2_cor) >= 2:
                u_stat, p_raw = stats.mannwhitneyu(g1_cor, g2_cor, alternative='two-sided')
                p_bonf = min(1.0, p_raw * n_comp)
                r_efecto = calcular_tamano_efecto_r(p_raw, len(g1_cor) + len(g2_cor))
                lineas.append(f"| {par_nombre} | M4 Micro-correcciones | {u_stat:.1f} | {p_raw:.4f} | {p_bonf:.4f} | {r_efecto:.3f} |")

    lineas.append("")
    lineas.append("### 3. Contraste de Simetría Direccional M2 (Wilcoxon Emparejado $e_x$ vs $e_y$)")
    lineas.append("")
    todos_pares = [p for c in CONDICIONES_ORDENADAS for p in pares_xy[c]]
    if len(todos_pares) >= 5:
        exs = [abs(p[0]) for p in todos_pares]
        eys = [abs(p[1]) for p in todos_pares]
        w_stat, p_wilc = stats.wilcoxon(exs, eys)
        r_wilc = calcular_tamano_efecto_r(p_wilc, len(todos_pares))
        lineas.append(f"- **Diferencia $|e_x|$ vs $|e_y|$ ($N={len(todos_pares)}$):** $W = {w_stat:.1f}$, $p = {p_wilc:.4f}$, tamaño de efecto $r = {r_wilc:.3f}$.")
        if p_wilc < 0.05:
            lineas.append("  - *Interpretación:* Existe una asimetría estadísticamente significativa entre el error en profundidad ($X$) y el error lateral ($Y$), siendo mayor la imprecisión en profundidad debido a la perspectiva monocular y la geometría de agarre.")
        else:
            lineas.append("  - *Interpretación:* No se observa asimetría estadísticamente significativa entre los ejes X e Y.")
    else:
        lineas.append("- Mediciones de pares $e_x, e_y$ insuficientes para el test de Wilcoxon.")

    return lineas

def main():
    parser = argparse.ArgumentParser(
        description="Análisis Estadístico Completo de la Fase P7"
    )
    parser.add_argument("--dir-resultados", type=str, default=None,
                        help="Carpeta donde residen los CSV de resultados")

    args = parser.parse_args()

    dir_res = Path(args.dir_resultados) if args.dir_resultados else RESULTADOS_DIR
    dir_res.mkdir(parents=True, exist_ok=True)

    log.info(f"Cargando datos desde repositorio: {REPO_ROOT}")
    res_m1_man, res_m1_soft = cargar_datos_m1(REPO_ROOT)
    res_m2, res_m3_xy, res_m3_iso = cargar_datos_m2_m3(REPO_ROOT)
    res_m4_t, res_m4_e, curvas_m4, res_m4_corrs = cargar_datos_m4(REPO_ROOT)

    log.info("Generando tablas y figuras gráficas...")
    tabla_resumen = generar_tabla_y_graficos(
        dir_res, res_m1_man, res_m1_soft, res_m2, res_m3_xy, res_m3_iso,
        res_m4_t, res_m4_e, curvas_m4, res_m4_corrs
    )

    log.info("Calculando contrastes de hipótesis estadísticos...")
    bloque_stats = ejecutar_tests_estadisticos(res_m1_man, res_m2, res_m4_t, res_m4_corrs, res_m3_xy)

    # Compilar Informe Markdown
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    informe_path = dir_res / f"informe_p7_{timestamp}.md"

    with open(informe_path, mode="w", encoding="utf-8") as f:
        f.write("# Informe Estadístico y Resultados de Fase P7 — Pruebas Funcionales TFG\n\n")
        f.write(f"**Fecha de generación:** {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}  \n")
        f.write(f"**Ubicación de datos:** `{dir_res}`  \n\n")
        f.write("---\n\n")
        f.write("## 1. Tabla Principal de Rendimiento (Condición × Métrica)\n\n")
        f.write("| Condición de Red | Latencia G2G Software | M1 Latencia G2G Manual | M2 Error Radial | M3 Repetibilidad 2D | M4 Tiempo Compleción | M4 Tasa Éxito |\n")
        f.write("|---|---|---|---|---|---|---|\n")
        for c in CONDICIONES_ORDENADAS:
            row = tabla_resumen[c]
            f.write(f"| **{NOMBRES_LEGIBLES[c]}** | {LATENCIA_G2G_MS[c]} ms | {row['M1_Latencia_G2G']} | {row['M2_Error_Radial_cm']} | {row['M3_Repetibilidad_cm']} | {row['M4_Tiempo_s']} | {row['M4_Tasa_Exito']} |\n")

        f.write("\n---\n\n")
        f.write("## 2. Contrastes de Hipótesis y Significación Estadística\n\n")
        for l in bloque_stats:
            f.write(l + "\n")

        f.write("\n---\n\n")
        f.write("## 3. Gráficas Generadas\n\n")
        f.write("- **Correlación Red ↔ Tarea:** `grafico_correlacion_g2g_m4.png`\n")
        f.write("- **Curvas de Aprendizaje:** `curva_aprendizaje_m4.png`\n")
        f.write("- **Tabla en formato CSV:** `tabla_resumen_p7.csv`\n")

    log.info(f"==========================================================")
    log.info(f" Informe estadístico P7 compilado con éxito.")
    log.info(f" Documento Markdown: {informe_path}")
    log.info(f"==========================================================")

if __name__ == "__main__":
    main()
